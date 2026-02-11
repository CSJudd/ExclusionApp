from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime


def write_audit_workbook(
    output_path,
    client_name,
    month,
    results,
    metadata
):
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create result sheets
    create_sheet(wb, "Staff Results", results.get("staff", []))
    create_sheet(wb, "Board Results", results.get("board", []))
    create_sheet(wb, "Vendor Results", results.get("vendors", []))

    # Consolidated Possible Matches
    possible_rows = collect_possible(results)
    create_sheet(wb, "Possible Matches", possible_rows)

    # Metadata sheet
    create_metadata_sheet(wb, metadata)

    wb.save(output_path)


def create_sheet(wb, title, rows):
    ws = wb.create_sheet(title)

    if not rows:
        ws.append(["No records"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(col, "") for col in headers])


def collect_possible(results):
    possible = []

    for category in results.values():
        for row in category:
            if (
                row.get("oig_status") == "POSSIBLE"
                or row.get("sam_status") == "POSSIBLE"
            ):
                possible.append(row)

    return possible


def create_metadata_sheet(wb, metadata):
    ws = wb.create_sheet("Run Metadata")

    ws.append(["Key", "Value"])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for key, value in metadata.items():
        ws.append([key, value])